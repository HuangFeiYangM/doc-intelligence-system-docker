"""
Tests for TableGenerator module.
"""
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.services.table_generator import TableGenerator, TableGeneratorError


class TestTableGenerator:
    """Test cases for TableGenerator."""

    @pytest.fixture
    def generator(self):
        """Create a TableGenerator instance."""
        return TableGenerator()

    @pytest.fixture
    def sample_template(self, temp_dir):
        """Create a sample Excel template."""
        template_path = temp_dir / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        # Add headers in first row
        ws["A1"] = "Name"
        ws["B1"] = "Date"
        ws["C1"] = "Amount"

        wb.save(template_path)
        return template_path

    def test_create_new(self, generator):
        """Test creating a new workbook."""
        generator.create_new()
        assert generator.workbook is not None

    def test_load_template(self, generator, sample_template):
        """Test loading an existing template."""
        generator.load_template(sample_template)
        assert generator.workbook is not None
        assert generator.template_path == sample_template

    def test_load_template_not_found(self, generator):
        """Test loading non-existent template."""
        with pytest.raises(TableGeneratorError) as exc_info:
            generator.load_template("/nonexistent/template.xlsx")
        assert "Template not found" in str(exc_info.value)

    def test_fill_data_with_mapping(self, generator, sample_template, temp_dir):
        """Test filling data with explicit field mapping."""
        output_path = temp_dir / "output.xlsx"

        field_mapping = {
            "Name": "A2",
            "Date": "B2",
            "Amount": "C2",
        }

        data = {
            "Name": "Test Project",
            "Date": "2024-03-15",
            "Amount": 10000,
        }

        generator.load_template(sample_template)
        generator.fill_data(data, field_mapping)
        generator.save(output_path)

        # Verify output
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws["A2"].value == "Test Project"
        assert ws["B2"].value == "2024-03-15"
        assert ws["C2"].value == 10000

    def test_fill_data_auto_fill(self, generator, temp_dir):
        """Test auto-filling data by matching headers."""
        # Create template with headers
        template_path = temp_dir / "auto_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "合同编号"
        ws["B1"] = "甲方"
        ws["C1"] = "金额"
        wb.save(template_path)

        output_path = temp_dir / "output_auto.xlsx"

        data = {
            "合同编号": "HT-001",
            "甲方": "ABC公司",
            "金额": 50000,
        }

        generator.load_template(template_path)
        generator.fill_data(data)  # No field mapping - auto-fill
        generator.save(output_path)

        # Verify output
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws["A2"].value == "HT-001"
        assert ws["B2"].value == "ABC公司"
        assert ws["C2"].value == 50000

    def test_fill_data_without_loading(self, generator):
        """Test filling data without loading template first."""
        with pytest.raises(TableGeneratorError) as exc_info:
            generator.fill_data({"field": "value"})
        assert "No template loaded" in str(exc_info.value)

    def test_save_without_loading(self, generator, temp_dir):
        """Test saving without loading workbook."""
        with pytest.raises(TableGeneratorError) as exc_info:
            generator.save(temp_dir / "test.xlsx")
        assert "No workbook to save" in str(exc_info.value)

    def test_generate_from_template_one_step(self, generator, sample_template, temp_dir):
        """Test one-step generation from template."""
        output_path = temp_dir / "one_step_output.xlsx"

        field_mapping = {
            "Name": "A2",
            "Date": "B2",
            "Amount": "C2",
        }

        data = {
            "Name": "Project X",
            "Date": "2024-12-01",
            "Amount": 99999,
        }

        result_path = generator.generate_from_template(
            sample_template, data, field_mapping, output_path
        )

        assert result_path == output_path
        assert output_path.exists()

    def test_create_template_static(self, temp_dir):
        """Test creating a new template."""
        template_path = temp_dir / "new_template.xlsx"
        headers = ["Field1", "Field2", "Field3"]

        result_path = TableGenerator.create_template(template_path, headers)

        assert result_path.exists()

        # Verify content
        wb = load_workbook(result_path)
        ws = wb.active
        assert ws["A1"].value == "Field1"
        assert ws["B1"].value == "Field2"
        assert ws["C1"].value == "Field3"

    def test_create_template_with_mapping(self, temp_dir):
        """Test creating template with field mapping."""
        template_path = temp_dir / "mapped_template.xlsx"
        headers = ["Name", "Date", "Amount"]
        field_mapping = {
            "Name": "B2",
            "Date": "C2",
            "Amount": "D2",
        }

        TableGenerator.create_template(template_path, headers, field_mapping)

        # Verify placeholders
        wb = load_workbook(template_path)
        ws = wb.active
        assert ws["B2"].value == "<Name>"
        assert ws["C2"].value == "<Date>"
        assert ws["D2"].value == "<Amount>"


class TestTableGeneratorEdgeCases:
    """Edge case tests for TableGenerator."""

    def test_fill_empty_data(self, temp_dir):
        """Test filling empty data."""
        template_path = temp_dir / "template.xlsx"
        output_path = temp_dir / "output.xlsx"

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Field"
        wb.save(template_path)

        generator = TableGenerator()
        generator.load_template(template_path)
        generator.fill_data({})
        generator.save(output_path)

        # Should not raise error
        assert output_path.exists()

    def test_fill_partial_data(self, temp_dir):
        """Test filling partial data (some fields missing)."""
        template_path = temp_dir / "template.xlsx"
        output_path = temp_dir / "output.xlsx"

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Field1"
        ws["B1"] = "Field2"
        wb.save(template_path)

        mapping = {"Field1": "A2", "Field2": "B2"}
        data = {"Field1": "Value1"}  # Field2 is missing

        generator = TableGenerator()
        generator.load_template(template_path)
        generator.fill_data(data, mapping)
        generator.save(output_path)

        # Verify partial fill
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws["A2"].value == "Value1"
        assert ws["B2"].value is None

    def test_multiple_rows_auto_fill(self, temp_dir):
        """Test auto-filling multiple rows."""
        template_path = temp_dir / "template.xlsx"
        output_path = temp_dir / "output.xlsx"

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Existing"
        ws["B2"] = "Data"
        wb.save(template_path)

        data = {"Name": "New", "Value": 123}

        generator = TableGenerator()
        generator.load_template(template_path)
        generator.fill_data(data)  # Should fill row 3
        generator.save(output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws["A2"].value == "Existing"  # Original data preserved
        assert ws["A3"].value == "New"  # New data in next row