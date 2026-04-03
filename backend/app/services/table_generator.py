"""
Table generator module for filling data into Excel templates.
"""
import logging
import os
from pathlib import Path
from typing import Any, Dict, Optional, Union

import openpyxl
from openpyxl import Workbook, load_workbook

from app.config import get_settings

# Configure logger
logger = logging.getLogger(__name__)

settings = get_settings()


class TableGeneratorError(Exception):
    """Exception raised for table generation errors."""
    pass


class TableGenerator:
    """Generator for filling data into Excel templates."""

    def __init__(self, template_path: Optional[Union[str, Path]] = None):
        """Initialize the table generator.

        Args:
            template_path: Path to the Excel template file
        """
        self.template_path = Path(template_path) if template_path else None
        self.workbook: Optional[Workbook] = None
        logger.debug(f"TableGenerator initialized with template: {template_path}")

    def load_template(self, template_path: Union[str, Path]) -> "TableGenerator":
        """Load an Excel template.

        Args:
            template_path: Path to the template file

        Returns:
            Self for method chaining

        Raises:
            TableGeneratorError: If template cannot be loaded
        """
        template_path = Path(template_path)
        logger.info(f"Loading template: {template_path}")

        if not template_path.exists():
            logger.error(f"Template not found: {template_path}")
            raise TableGeneratorError(f"Template not found: {template_path}")

        try:
            self.template_path = template_path
            self.workbook = load_workbook(template_path)
            logger.info(f"Template loaded successfully: {template_path}")
            return self
        except Exception as e:
            logger.error(f"Failed to load template {template_path}: {e}", exc_info=True)
            raise TableGeneratorError(f"Failed to load template: {str(e)}")

    def create_new(self) -> "TableGenerator":
        """Create a new blank workbook.

        Returns:
            Self for method chaining
        """
        self.workbook = Workbook()
        logger.debug("Created new blank workbook")
        return self

    def fill_data(
        self,
        data: Dict[str, Any],
        field_mapping: Optional[Dict[str, str]] = None,
        sheet_name: Optional[str] = None
    ) -> "TableGenerator":
        """Fill data into the template.

        Args:
            data: Dictionary with field names and values
            field_mapping: Mapping of field names to cell addresses (e.g., {"name": "B2"})
            sheet_name: Target sheet name (uses active sheet if None)

        Returns:
            Self for method chaining

        Raises:
            TableGeneratorError: If data cannot be filled
        """
        if not self.workbook:
            raise TableGeneratorError("No template loaded. Call load_template() first.")

        try:
            if sheet_name:
                try:
                    sheet = self.workbook[sheet_name]
                except KeyError:
                    logger.error(f"Sheet not found: {sheet_name}")
                    raise TableGeneratorError(f"Sheet not found: {sheet_name}")
            else:
                sheet = self.workbook.active

            filled_count = 0
            if field_mapping:
                # Use explicit field mapping
                for field_name, cell_address in field_mapping.items():
                    if field_name in data:
                        sheet[cell_address] = data[field_name]
                        filled_count += 1
                    else:
                        logger.warning(f"Field '{field_name}' not found in data")
            else:
                # Auto-fill: assume first row is headers, find matching columns
                filled_count = self._auto_fill_data(sheet, data)

            logger.info(f"Filled {filled_count} fields into template")
            return self
        except TableGeneratorError:
            raise
        except Exception as e:
            logger.error(f"Failed to fill data: {e}", exc_info=True)
            raise TableGeneratorError(f"Failed to fill data: {str(e)}")

    def _auto_fill_data(self, sheet, data: Dict[str, Any]) -> int:
        """Auto-fill data by matching headers to field names.

        Args:
            sheet: Worksheet to fill
            data: Data dictionary

        Returns:
            Number of fields filled
        """
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        if not headers:
            logger.warning("No headers found in first row")
            return 0

        # Find first empty row or use second row
        target_row = 2
        for row_idx in range(2, sheet.max_row + 2):
            if all(sheet.cell(row=row_idx, column=col).value is None
                   for col in range(1, sheet.max_column + 1)):
                target_row = row_idx
                break

        # Fill data
        filled_count = 0
        for field_name, value in data.items():
            if field_name in headers:
                col_idx = headers[field_name]
                sheet.cell(row=target_row, column=col_idx, value=value)
                filled_count += 1
            else:
                logger.debug(f"Field '{field_name}' not matched to any header")

        return filled_count

    def save(self, output_path: Union[str, Path]) -> Path:
        """Save the filled workbook.

        Args:
            output_path: Path to save the file

        Returns:
            Path to the saved file

        Raises:
            TableGeneratorError: If save fails
        """
        if not self.workbook:
            raise TableGeneratorError("No workbook to save")

        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(output_path)
            logger.info(f"Workbook saved to: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}", exc_info=True)
            raise TableGeneratorError(f"Failed to save workbook: {str(e)}")

    def generate_from_template(
        self,
        template_path: Union[str, Path],
        data: Dict[str, Any],
        field_mapping: Dict[str, str],
        output_path: Union[str, Path]
    ) -> Path:
        """One-step method to generate Excel from template.

        Args:
            template_path: Path to template file
            data: Data to fill
            field_mapping: Field to cell mapping
            output_path: Output file path

        Returns:
            Path to generated file
        """
        logger.info(f"Generating Excel from template: {template_path}")
        try:
            result = (
                self.load_template(template_path)
                .fill_data(data, field_mapping)
                .save(output_path)
            )
            logger.info(f"Excel generation completed: {output_path}")
            return result
        except Exception as e:
            logger.error(f"Excel generation failed: {e}", exc_info=True)
            raise

    @staticmethod
    def create_template(
        output_path: Union[str, Path],
        headers: list,
        field_mapping: Optional[Dict[str, str]] = None
    ) -> Path:
        """Create a new Excel template with headers.

        Args:
            output_path: Path to save the template
            headers: List of column headers
            field_mapping: Optional mapping of fields to cells

        Returns:
            Path to created template
        """
        logger.info(f"Creating template with headers: {headers}")

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Template"

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)

        # If field mapping provided, add sample data row
        if field_mapping:
            for field, cell_addr in field_mapping.items():
                sheet[cell_addr] = f"<{field}>"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Template created: {output_path}")
        return output_path
